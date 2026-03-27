from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Count
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from functools import wraps
import json, io, csv
from datetime import datetime, date

from .models import (Hardware, Employee, TrashHardware, HardwareProperty,
                     CommandLog, CustomUser, HardwareType, ResignationRequest,
                     TransferRequest, HardwareApproval, Task, Message, STATUS_CHOICES)

# ─── HELPERS ──────────────────────────────────────────────────────────────────

def get_hw_types():
    types = list(HardwareType.objects.filter(is_active=True).values('name','icon'))
    if not types:
        defaults = [('Laptop','💻'),('Desktop','🖥'),('CPU','🔧'),('Server','🗄'),
                    ('Monitor','📺'),('Camera','📷'),('Printer','🖨'),('Scanner','📠'),
                    ('Mouse','🖱'),('Keyboard','⌨'),('UPS','🔋'),('Switch','🔀'),('Router','📡'),('Other','📦')]
        for i,(n,ic) in enumerate(defaults):
            HardwareType.objects.get_or_create(name=n, defaults={'icon':ic,'order':i})
        types = list(HardwareType.objects.filter(is_active=True).values('name','icon'))
    return types

def get_location_filtered_hardware(user, qs=None):
    if qs is None:
        qs = Hardware.objects.select_related('assigned_to')
    if user.has_location_filter:
        qs = qs.filter(location__iexact=user.location)
    return qs

def user_can_access_hw(user, hw):
    if not user.has_location_filter:
        return True
    return hw.location.lower() == user.location.lower()

def editor_required(f):
    @wraps(f)
    def wrap(request, *args, **kwargs):
        if not request.user.is_authenticated: return redirect('/login/')
        if not request.user.can_edit: return render(request, 'hardware/403.html', status=403)
        return f(request, *args, **kwargs)
    return wrap

def admin_required(f):
    @wraps(f)
    def wrap(request, *args, **kwargs):
        if not request.user.is_authenticated: return redirect('/login/')
        if not request.user.can_admin: return render(request, 'hardware/403.html', status=403)
        return f(request, *args, **kwargs)
    return wrap

def superadmin_required(f):
    @wraps(f)
    def wrap(request, *args, **kwargs):
        if not request.user.is_authenticated: return redirect('/login/')
        if not request.user.is_superadmin_role: return render(request, 'hardware/403.html', status=403)
        return f(request, *args, **kwargs)
    return wrap

# ─── AUTH ─────────────────────────────────────────────────────────────────────

def login_view(request):
    if request.user.is_authenticated: return redirect('/')
    error = None
    if request.method == 'POST':
        user = authenticate(request, username=request.POST.get('username'), password=request.POST.get('password'))
        if user:
            login(request, user)
            return redirect(request.GET.get('next', '/'))
        error = 'Invalid credentials. Access denied.'
    return render(request, 'hardware/login.html', {'error': error})

def logout_view(request):
    logout(request)
    return redirect('/login/')

# ─── RESIGNATION (PUBLIC — no login needed) ───────────────────────────────────

def resignation_page(request):
    """Public page for employees to submit resignation"""
    if request.method == 'POST':
        try:
            emp_id = request.POST.get('emp_id','').strip()
            emp_name = request.POST.get('employee_name','').strip()
            reason = request.POST.get('reason','').strip()
            letter = request.POST.get('resignation_letter','').strip()
            last_date = request.POST.get('last_working_date','') or None

            if not emp_id or not emp_name or not reason:
                return JsonResponse({'success': False, 'error': 'All fields are required'})

            # Find employee location for routing to correct superadmin
            location = ''
            try:
                emp = Employee.objects.get(emp_id=emp_id)
                location = emp.location
            except Employee.DoesNotExist:
                pass

            res_obj = ResignationRequest.objects.create(
                emp_id=emp_id,
                employee_name=emp_name,
                reason=reason,
                resignation_letter=letter,
                last_working_date=last_date or None,
                location=location,
            )
            # Notify area admins
            send_message_to_area_admins(
                location=location,
                title=f'New Resignation: {emp_name} ({emp_id})',
                body=f'Employee {emp_name} has submitted a resignation.\nReason: {reason}\nLast Working Date: {last_date or "Not specified"}',
                msg_type='resignation',
                sender_name=emp_name,
                related_id=res_obj.pk,
            )
            return JsonResponse({'success': True, 'message': 'Resignation submitted successfully. Admin will review it.'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return render(request, 'hardware/resignation.html')

def hardware_report_page(request):
    """Public — employee reports hardware issue. Also creates HardwareApproval."""
    if request.method == 'POST':
        try:
            emp_id = request.POST.get('emp_id','').strip()
            hw_id = request.POST.get('hw_id','').strip()
            issue = request.POST.get('issue','').strip()
            if not emp_id or not issue:
                return JsonResponse({'success': False, 'error': 'All fields required'})
            location = ''; emp_name = emp_id
            try:
                emp = Employee.objects.get(emp_id=emp_id)
                location = emp.location; emp_name = emp.name
            except Employee.DoesNotExist: pass
            hw = None
            if hw_id:
                try:
                    hw = Hardware.objects.get(hw_id=hw_id)
                    hw.notes = hw.notes + f"\n[REPORT by {emp_id} on {date.today()}]: {issue}"
                    hw.status = 'maintenance'
                    hw.save()
                except Hardware.DoesNotExist: pass
            # Create HardwareApproval so it appears in Approvals tab
            HardwareApproval.objects.create(
                emp_id=emp_id, employee_name=emp_name,
                hardware=hw, hw_id_text=hw_id,
                request_type='repair', issue_description=issue,
                location=location,
            )
            return JsonResponse({'success': True, 'message': 'Report submitted. IT team will be notified.'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return render(request, 'hardware/hardware_report.html')

@login_required
def dashboard(request):
    hw_qs = get_location_filtered_hardware(request.user)
    total_hw = hw_qs.count()
    active_hw = hw_qs.filter(status='active').count()
    total_emp = Employee.objects.filter(status='active').count()
    total_trash = TrashHardware.objects.count()
    hw_types = get_hw_types()
    hw_by_type = {t['name']: hw_qs.filter(hardware_type=t['name']).count() for t in hw_types}
    recent_hw = hw_qs.order_by('-created_at')[:6]
    recent_trash = TrashHardware.objects.order_by('-disposed_date')[:4]
    pending_resignations = ResignationRequest.objects.filter(status='pending')
    pending_transfers = TransferRequest.objects.filter(status='pending')
    pending_hw_approvals = HardwareApproval.objects.filter(status='pending')
    if request.user.has_location_filter:
        pending_resignations = pending_resignations.filter(location=request.user.location)
        pending_transfers = pending_transfers.filter(from_location=request.user.location)
        pending_hw_approvals = pending_hw_approvals.filter(location=request.user.location)
    all_locations = Hardware.objects.values_list('location',flat=True).distinct().exclude(location='').order_by('location')
    return render(request, 'hardware/dashboard.html', {
        'total_hw':total_hw,'total_emp':total_emp,'total_trash':total_trash,
        'active_hw':active_hw,'hw_by_type':hw_by_type,'hw_types':hw_types,
        'recent_hw':recent_hw,'recent_trash':recent_trash,
        'pending_resignations':pending_resignations.count(),
        'pending_transfers':pending_transfers.count(),
        'pending_hw_approvals':pending_hw_approvals.count(),
        'total_pending': pending_resignations.count() + pending_transfers.count() + pending_hw_approvals.count(),
        'all_locations':all_locations,
    })

# ─── HARDWARE ─────────────────────────────────────────────────────────────────

@login_required
def hardware_list(request):
    hw_type = request.GET.get('type','')
    status = request.GET.get('status','')
    search = request.GET.get('search','')
    loc = request.GET.get('location','')
    hardware = get_location_filtered_hardware(request.user)
    if hw_type: hardware = hardware.filter(hardware_type=hw_type)
    if status: hardware = hardware.filter(status=status)
    if loc: hardware = hardware.filter(location__icontains=loc)
    if search:
        hardware = hardware.filter(Q(hw_id__icontains=search)|Q(brand__icontains=search)|Q(model_name__icontains=search)|Q(serial_number__icontains=search)|Q(location__icontains=search))
    hardware = hardware.order_by('-created_at')
    all_locations = Hardware.objects.values_list('location',flat=True).distinct().exclude(location='').order_by('location')
    return render(request,'hardware/hardware_list.html',{
        'hardware':hardware,'hardware_types':get_hw_types(),'status_choices':STATUS_CHOICES,
        'selected_type':hw_type,'selected_status':status,'search':search,
        'all_locations':all_locations,'selected_location':loc,
    })

@login_required
def hardware_detail(request, pk):
    hw = get_object_or_404(Hardware, pk=pk)
    if not user_can_access_hw(request.user, hw): return render(request,'hardware/403.html',status=403)
    return render(request,'hardware/hardware_detail.html',{'hw':hw,'props':hw.properties.all(),'status_choices':STATUS_CHOICES})

@login_required
def hardware_status_change(request, pk):
    if not request.user.can_edit: return JsonResponse({'success':False,'error':'Insufficient privileges'})
    if request.method == 'POST':
        hw = get_object_or_404(Hardware, pk=pk)
        if not user_can_access_hw(request.user, hw): return JsonResponse({'success':False,'error':'Location restricted'})
        new_status = request.POST.get('status')
        if new_status not in [s for s,_ in STATUS_CHOICES]: return JsonResponse({'success':False,'error':'Invalid status'})
        hw.status = new_status; hw.save()
        return JsonResponse({'success':True,'status':hw.status})
    return JsonResponse({'success':False})

@login_required
@login_required
def hardware_search_api(request):
    """Search unassigned hardware. Empty/dot/all = return all unassigned"""
    q = request.GET.get('q','').strip()
    limit = int(request.GET.get('limit', 30))
    hardware = Hardware.objects.filter(assigned_to__isnull=True)
    if q and q not in ['.','all','']:
        hardware = hardware.filter(
            Q(hw_id__icontains=q)|Q(brand__icontains=q)|
            Q(model_name__icontains=q)|Q(hardware_type__icontains=q)|
            Q(serial_number__icontains=q)
        )
    hardware = hardware.order_by('hardware_type','brand')[:limit]
    results = [{'id':hw.pk,'hw_id':hw.hw_id,'type':hw.hardware_type,'brand':hw.brand,'model':hw.model_name,'serial':hw.serial_number} for hw in hardware]
    return JsonResponse({'results':results})

@editor_required
def hardware_add(request):
    if request.method == 'POST':
        try:
            location = request.POST.get('location','')
            if request.user.has_location_filter: location = request.user.location
            hw = Hardware(
                hw_id=request.POST.get('hw_id'),
                hardware_type=request.POST.get('hardware_type'),
                brand=request.POST.get('brand'),
                model_name=request.POST.get('model_name'),
                serial_number=request.POST.get('serial_number'),
                purchase_date=request.POST.get('purchase_date'),
                price=request.POST.get('price') or 0,
                status=request.POST.get('status','active'),
                location=location,
                specifications=request.POST.get('specifications',''),
                notes=request.POST.get('notes',''),
                created_by=request.user,
            )
            if request.POST.get('warranty_expiry'): hw.warranty_expiry = request.POST.get('warranty_expiry')
            assigned = request.POST.get('assigned_to')
            if assigned: hw.assigned_to_id = assigned
            hw.save()
            for i,(k,v) in enumerate(zip(request.POST.getlist('prop_key[]'),request.POST.getlist('prop_val[]'))):
                if k.strip() and v.strip(): HardwareProperty.objects.create(hardware=hw,key=k.strip(),value=v.strip(),order=i)
            return JsonResponse({'success':True,'id':hw.pk,'hw_id':hw.hw_id})
        except Exception as e:
            return JsonResponse({'success':False,'error':str(e)})
    all_locations = Hardware.objects.values_list('location',flat=True).distinct().exclude(location='').order_by('location')
    return render(request,'hardware/hardware_add.html',{
        'hardware_types':get_hw_types(),'status_choices':STATUS_CHOICES,
        'employees':Employee.objects.filter(status='active'),
        'all_locations':all_locations,'user_location':request.user.location,
    })

@editor_required
def hardware_edit(request, pk):
    hw = get_object_or_404(Hardware, pk=pk)
    if not user_can_access_hw(request.user, hw): return render(request,'hardware/403.html',status=403)
    if request.method == 'POST':
        try:
            hw.hw_id=request.POST.get('hw_id',hw.hw_id)
            hw.hardware_type=request.POST.get('hardware_type',hw.hardware_type)
            hw.brand=request.POST.get('brand',hw.brand)
            hw.model_name=request.POST.get('model_name',hw.model_name)
            hw.serial_number=request.POST.get('serial_number',hw.serial_number)
            hw.purchase_date=request.POST.get('purchase_date',hw.purchase_date)
            hw.price=request.POST.get('price',hw.price)
            hw.status=request.POST.get('status',hw.status)
            hw.specifications=request.POST.get('specifications',hw.specifications)
            hw.notes=request.POST.get('notes',hw.notes)
            if request.user.can_admin: hw.location=request.POST.get('location',hw.location)
            if request.POST.get('warranty_expiry'): hw.warranty_expiry=request.POST.get('warranty_expiry')
            assigned=request.POST.get('assigned_to')
            hw.assigned_to_id=assigned if assigned else None
            hw.save()
            hw.properties.all().delete()
            for i,(k,v) in enumerate(zip(request.POST.getlist('prop_key[]'),request.POST.getlist('prop_val[]'))):
                if k.strip() and v.strip(): HardwareProperty.objects.create(hardware=hw,key=k.strip(),value=v.strip(),order=i)
            return JsonResponse({'success':True})
        except Exception as e:
            return JsonResponse({'success':False,'error':str(e)})
    all_locations = Hardware.objects.values_list('location',flat=True).distinct().exclude(location='').order_by('location')
    return render(request,'hardware/hardware_edit.html',{
        'hw':hw,'hardware_types':get_hw_types(),'status_choices':STATUS_CHOICES,
        'employees':Employee.objects.filter(status='active'),
        'props':list(hw.properties.values('key','value')),'all_locations':all_locations,
    })

# ─── HARDWARE TYPES MANAGEMENT ────────────────────────────────────────────────

@superadmin_required
def hardware_types_manage(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add':
            try:
                ht, created = HardwareType.objects.get_or_create(
                    name=request.POST.get('name','').strip(),
                    defaults={'icon':request.POST.get('icon','📦'),'order':HardwareType.objects.count()}
                )
                if not created:
                    ht.icon = request.POST.get('icon','📦')
                    ht.is_active = True
                    ht.save()
                return JsonResponse({'success':True,'id':ht.pk,'name':ht.name,'icon':ht.icon})
            except Exception as e:
                return JsonResponse({'success':False,'error':str(e)})
        elif action == 'edit':
            try:
                ht = get_object_or_404(HardwareType, pk=request.POST.get('id'))
                ht.name = request.POST.get('name', ht.name)
                ht.icon = request.POST.get('icon', ht.icon)
                ht.save()
                return JsonResponse({'success':True})
            except Exception as e:
                return JsonResponse({'success':False,'error':str(e)})
        elif action == 'delete':
            try:
                ht = get_object_or_404(HardwareType, pk=request.POST.get('id'))
                ht.is_active = not ht.is_active  # TOGGLE
                ht.save()
                return JsonResponse({'success':True, 'is_active': ht.is_active})
            except Exception as e:
                return JsonResponse({'success':False,'error':str(e)})
        elif action == 'save_fields':
            try:
                ht = get_object_or_404(HardwareType, pk=request.POST.get('id'))
                ht.custom_fields = request.POST.get('fields','')
                ht.save()
                return JsonResponse({'success':True})
            except Exception as e:
                return JsonResponse({'success':False,'error':str(e)})
    types = HardwareType.objects.all()
    return render(request,'hardware/hardware_types.html',{'types':types})

# ─── EXCEL IMPORT / EXPORT ────────────────────────────────────────────────────

@login_required
def export_hardware_excel(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse("openpyxl not installed.", status=500)
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Hardware Inventory"
    headers = ['HW ID','Type','Brand','Model','Serial No','Status','Location','Purchase Date','Warranty Expiry','Price','Assigned To (Emp ID)','Specifications','Notes']
    hf = PatternFill("solid", fgColor="0F1729")
    for col,h in enumerate(headers,1):
        cell = ws.cell(row=1,column=col,value=h)
        cell.fill=hf; cell.font=Font(bold=True,color="00F5FF",name="Consolas")
        cell.alignment=Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width=max(15,len(h)+2)
    for hw in get_location_filtered_hardware(request.user).select_related('assigned_to').all():
        ws.append([hw.hw_id,hw.hardware_type,hw.brand,hw.model_name,hw.serial_number,hw.status,hw.location,str(hw.purchase_date),str(hw.warranty_expiry) if hw.warranty_expiry else '',float(hw.price),hw.assigned_to.emp_id if hw.assigned_to else '',hw.specifications,hw.notes])
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    r=HttpResponse(buf.read(),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    r['Content-Disposition']=f'attachment; filename="DB_Hardware_{date.today()}.xlsx"'
    return r

@editor_required
def import_hardware_excel(request):
    if request.method == "POST":
        try:
            import openpyxl
        except ImportError:
            return JsonResponse({"success": False, "error": "openpyxl not installed"})
        f = request.FILES.get("excel_file")
        if not f:
            return JsonResponse({"success": False, "error": "No file uploaded"})
        try:
            wb = openpyxl.load_workbook(f, data_only=True)
            ws = wb.active
            headers = [str(ws.cell(1, c).value or "").strip().lower() for c in range(1, ws.max_column+1)]
            added = 0; skipped = 0; errors = []

            def safe_str(val):
                if val is None: return ""
                return str(val).strip()

            def safe_date(val):
                if val is None: return date.today()
                if hasattr(val, "date"): return val.date()
                if hasattr(val, "year"): return val
                s = str(val).strip()
                if not s or s.lower() in ["none","nan",""]: return date.today()
                for fmt in ["%Y-%m-%d","%d-%m-%Y","%d/%m/%Y","%m/%d/%Y","%d-%b-%Y"]:
                    try: return datetime.strptime(s, fmt).date()
                    except: pass
                return date.today()

            def safe_float(val):
                if val is None: return 0
                try: return float(val)
                except: return 0

            def get_col(row, name, default=""):
                try:
                    if name in headers:
                        idx = headers.index(name)
                        return row[idx] if idx < len(row) else default
                except: pass
                return default

            hw_type_post = request.POST.get("import_type", "")

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row): continue
                if all(v is None or str(v).strip() == "" for v in row): continue
                try:
                    first_val = safe_str(row[0])
                    if not first_val: continue

                    is_sno = False
                    try:
                        int(float(first_val))
                        is_sno = True
                    except: pass

                    if is_sno:
                        # ── Detect type from sheet name first, then headers ──
                        sheet_name = ws.title.strip().lower()
                        h_str = " ".join(headers)

                        if hw_type_post:
                            detected_type = hw_type_post
                        elif "desktop" in sheet_name or "dt / tc" in h_str or "dt/tc" in h_str or "hostname" in h_str:
                            detected_type = "Desktop"
                        elif "laptop" in sheet_name or "laptop" in h_str:
                            detected_type = "Laptop"
                        elif "printer" in sheet_name or "printer" in h_str:
                            detected_type = "Printer"
                        elif "cctv" in sheet_name or "dvr" in h_str or "nvr" in h_str:
                            detected_type = "CCTV"
                        elif "ill" in sheet_name or "isp" in h_str or "circuit" in h_str:
                            detected_type = "ILL/BB"
                        elif "processor" in h_str or "ram" in h_str:
                            detected_type = "Laptop"
                        else:
                            detected_type = "Other"

                        prefix_map = {
                            "Laptop": "LAPT", "Desktop": "DESK", "Printer": "PRIN",
                            "CCTV": "CCTV", "ILL/BB": "ILLB", "Server": "SERV",
                            "Scanner": "SCAN", "TV": "TV", "Mouse": "MOUS",
                            "Keyboard": "KEYB", "Switch": "SWCH", "Other": "OTHR",
                        }
                        prefix = prefix_map.get(detected_type, detected_type[:4].upper())

                        # ── Get REAL serial from CSV ──
                        serial_raw = ""
                        for sk in ["serial no", "serial no.", "serial number", "sr no", "sr. no", "serial"]:
                            v = get_col(row, sk)
                            if v and safe_str(v) not in ["", "None", "nan"]:
                                serial_raw = safe_str(v)
                                break

                        # ── Get real HW ID from SAP/Asset code ──
                        sap_raw = ""
                        for ak in ["sap asset code", "sap code(dt)", "sap code", "asset code"]:
                            v = get_col(row, ak)
                            if v and safe_str(v) not in ["", "None", "nan"]:
                                sap_raw = safe_str(v)
                                break

                        # Build hw_id
                        if sap_raw:
                            hw_id = f"{prefix}-{sap_raw}"
                        else:
                            hw_id = f"{prefix}-{row_idx:04d}"
                        # Ensure unique
                        base_hw_id = hw_id
                        counter = 1
                        while Hardware.objects.filter(hw_id=hw_id).exists():
                            hw_id = f"{base_hw_id}-{counter}"
                            counter += 1

                        # Build serial
                        if serial_raw:
                            serial = serial_raw
                        else:
                            serial = f"{prefix}-SN-{row_idx:04d}"
                        # Ensure unique serial
                        base_serial = serial
                        counter = 1
                        while Hardware.objects.filter(serial_number=serial).exists():
                            serial = f"{base_serial}-{counter}"
                            counter += 1

                        location = ""
                        for lk in ["user location", "center name", "location", "state", "asset location"]:
                            lv = get_col(row, lk)
                            if lv and safe_str(lv) not in ["", "None", "nan"]:
                                location = safe_str(lv)
                                break
                        if request.user.has_location_filter:
                            location = request.user.location

                        brand = safe_str(get_col(row,"make") or get_col(row,"brand") or get_col(row,"isp name") or "Imported")
                        model = safe_str(get_col(row,"model") or get_col(row,"suggested model") or "Imported")
                        price = safe_float(get_col(row,"price without gst") or get_col(row,"cost approx") or 0)

                        # Get employee
                        emp_id_val = safe_str(get_col(row,"emp id") or get_col(row,"empl id") or "")
                        assigned_emp = None
                        if emp_id_val:
                            try:
                                assigned_emp = Employee.objects.get(emp_id=emp_id_val)
                            except Employee.DoesNotExist:
                                pass

                        hw = Hardware.objects.create(
                            hw_id=hw_id, hardware_type=detected_type,
                            brand=brand or "Imported", model_name=model or "Imported",
                            serial_number=serial, purchase_date=date.today(),
                            price=price, status="active", location=location,
                            assigned_to=assigned_emp,
                            notes=f"Imported from Excel row {row_idx}",
                            created_by=request.user,
                        )
                        for col_idx, (header, val) in enumerate(zip(headers, row)):
                            if val is not None and str(val).strip() and header and header != "s.no":
                                HardwareProperty.objects.create(
                                    hardware=hw,
                                    key=header.replace("\n"," ").strip().title()[:100],
                                    value=safe_str(val)[:500],
                                    order=col_idx
                                )
                        added += 1

                    else:
                        # Standard DB Portal format
                        hw_id = first_val
                        if Hardware.objects.filter(hw_id=hw_id).exists():
                            skipped += 1; continue
                        location = safe_str(row[6]) if len(row) > 6 else ""
                        if request.user.has_location_filter: location = request.user.location
                        serial = safe_str(row[4]) if len(row) > 4 else hw_id + "-SN"
                        if not serial: serial = hw_id + "-SN"
                        if Hardware.objects.filter(serial_number=serial).exists():
                            serial = hw_id + "-SN-" + str(row_idx)
                        Hardware.objects.create(
                            hw_id=hw_id,
                            hardware_type=safe_str(row[1]) if len(row) > 1 else "Other",
                            brand=safe_str(row[2]) if len(row) > 2 else "",
                            model_name=safe_str(row[3]) if len(row) > 3 else "",
                            serial_number=serial,
                            status=safe_str(row[5]) if len(row) > 5 else "active",
                            location=location,
                            purchase_date=safe_date(row[7] if len(row) > 7 else None),
                            price=safe_float(row[9] if len(row) > 9 else 0),
                            specifications=safe_str(row[11]) if len(row) > 11 else "",
                            notes=safe_str(row[12]) if len(row) > 12 else "",
                            created_by=request.user,
                        )
                        added += 1

                except Exception as e:
                    errors.append(f"Row {row_idx}: {str(e)}")

            return JsonResponse({"success": True, "added": added, "skipped": skipped, "errors": errors[:5]})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})
    return JsonResponse({"success": False})

def export_template_excel(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse("openpyxl not installed.", status=500)

    hw_type = request.GET.get("type", "")

    TYPE_HEADERS = {
        "Laptop": ["S.No","State","Emp ID","EMP Name","Grade","Designation","Department","User Location","Region","Business Unit","Asset Location","Asset Type","Make","Model","Serial No","Processor Type","Processor Series","RAM","HDD Size","Hard Disk Type","Age in Year","OS","Generation","Current Hostname","Display Size","SAP Asset Code","PO No","PO Date","Price Without GST","Budget Year","Old Emp Id","Old EMP Name","Remarks"],
        "Desktop": ["S.No","State","Center Name","DT / TC","Department","Designation","Grade","Region","Division","OS","Make","Model","Serial No","SAP Code","HDD/SSD","HDD Size","RAM","Processor","Generation","DT Age","LCD Size (Inch)","LCD Model","LCD Serial No","LCD SAP Code","LCD Make","LCD Age","Current Status","User Working Details","Other Comments"],
        "Printer": ["S.No","State","Location","Press/Bureau/Office","Printer Requirement A4/A3/DMP","Suggested Model","Qty","Cost Approx","If Network - How Many A4 Free","Age","Reason (Justification)","Can Be Reused"],
        "CCTV": ["S.No","State","Center Name","Office/Press/Bureau","Surveillance Device Type","Total Ports DVR/NVR","Free Ports DVR/NVR","Using Ports","Camera VGA Non-IP Count","Camera IP 2MP Count","Camera IP >2MP Count","Camera Type","NVR/DVR Age","NVR/DVR Company","NVR/DVR Model","HDD Size","Backup Days","LCD Display","All Cameras Working","Non-Operational Since","Action Taken","Screenshot Link"],
        "ILL/BB": ["S.No","State","IT Engineer Location","Link Location","Office Type","ISP Name","Link Type","Link Status","Circuit ID","Fiber/RF/Copper","Speed MBPS","Yearly Cost","Link IP Address","Link Postal Address","PO No","PO Date","PO Bill Period Months","Next Renewal Date","Billing Cycle","Payment Location","Payment Done Upto","Payment Status","Amount Paid","Remark"],
        "Server": ["S.No","State","Location","Make","Model","Serial No","CPU","RAM","Storage RAID","OS","IP Address","Rack Unit","Network Ports","iDRAC IP","Power Supplies","Purchase Date","Warranty Expiry","Price"],
        "TV": ["S.No","State","Location","Make","Model","Serial No","Screen Size inch","Resolution","Panel Type","Smart TV","Ports","SAP Code","Installation Location","Purchase Date","Price"],
        "Scanner": ["S.No","State","Location","Make","Model","Serial No","Scanner Type","Resolution DPI","Scan Size","Connectivity","ADF Capacity","Purchase Date","Price"],
        "Switch": ["S.No","State","Location","Make","Model","Serial No","Number of Ports","Port Speed","Managed Unmanaged","IP Address","VLAN Support","PoE Support","Purchase Date","Price"],
        "Generic": ["HW ID","Type","Brand","Model","Serial No","Status","Location","Purchase Date","Warranty Expiry","Price","Assigned To Emp ID","Specifications","Notes"],
    }

    SAMPLE_ROWS = {
        "Laptop": [1,"Rajasthan","EMP-001","Rahul Sharma","L2","Developer","IT","Jaipur","State","DB Group","Head Office","Laptop","Dell","Inspiron 15 3520","DL-SN-001","Intel Core i5","12th Gen","8GB DDR4","512GB","SSD",2,"Windows 11","12th","SWASTIK-PC","15.6","SAP001","PO-001","2024-01-15",52000,"2024","","",""],
        "Desktop": [1,"MP","Head Office","Desktop","IT","Developer","L3","Corp","Editorial","Windows 11","Dell","OptiPlex 3000","SN-001","SAP-001","SSD","512GB","8GB","Intel Core i3-12100","12th",2,"21.5","LG 22MP410","LCD-SN","LCD-SAP","LG",1,"Active","Working fine",""],
        "Printer": [1,"Rajasthan","Jaipur","Office","Laser A4","HP LaserJet M208dw",1,18000,1,5,"For HR Department","No"],
        "CCTV": [1,"MP","Bhopal Press","Press","DVR+Camera",8,2,6,4,2,0,"IP",3,"Hikvision","DS-7208","2TB",15,"Yes","Yes","","",""],
        "ILL/BB": [1,"MP","Bhopal","Head Office","Corp Office","Airtel","ILL","Active","CKTID-001","Fiber",100,120000,"192.168.1.1","123 Main St","PO-001","2024-01-01",12,"2025-01-01","Annual","Corporate","2024-12-31","Paid",120000,""],
        "Server": [1,"MP","Server Room","Dell","PowerEdge R740","SRV-001","Dual Xeon","64GB ECC","4TB RAID-5","Ubuntu 22.04","192.168.1.10","2U","4x 1Gbps","192.168.1.11","Dual PSU","2022-01-15","2025-01-15",350000],
        "TV": [1,"MP","Conference Room","Samsung","UA55CU8000","TV-001",55,"4K UHD","LED","Yes","HDMI x3","SAP-TV-001","Conf Hall","2023-06-01",45000],
        "Scanner": [1,"MP","Head Office","Epson","Perfection V39","SC-001","Flatbed","4800","A4","USB",0,"2022-04-01",8500],
        "Switch": [1,"MP","Server Room","Cisco","SG300-28P","SW-001",28,"Gigabit","Managed","192.168.1.2","Yes","Yes","2021-09-01",25000],
        "Generic": ["HW-001","Laptop","Dell","Inspiron 15","SN12345","active","Head Office","2024-01-15","2026-01-15",55000,"EMP-001","Intel i5 8GB",""],
    }

    hf = PatternFill("solid", fgColor="0F1729")
    hfont = Font(bold=True, color="00F5FF", name="Consolas")
    center_align = Alignment(horizontal="center")

    def make_sheet(wb, title, headers, sample=None):
        ws = wb.create_sheet(title=title.replace("/","_")[:31])
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = hf
            cell.font = hfont
            cell.alignment = center_align
            ws.column_dimensions[cell.column_letter].width = max(16, len(str(h)) + 2)
        if sample:
            ws.append(sample)
        return ws

    if hw_type and hw_type in TYPE_HEADERS:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = hw_type.replace('/','_')[:31]
        headers = TYPE_HEADERS[hw_type]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = hf
            cell.font = hfont
            cell.alignment = center_align
            ws.column_dimensions[cell.column_letter].width = max(16, len(str(h)) + 2)
        sample = SAMPLE_ROWS.get(hw_type)
        if sample:
            ws.append(sample)
        filename = "DB_" + hw_type.replace("/","_") + "_Template.xlsx"
    else:
        wb = openpyxl.Workbook()
        first = True
        for t, headers in TYPE_HEADERS.items():
            if first:
                ws = wb.active
                ws.title = t.replace("/","_")[:31]
                for col, h in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=h)
                    cell.fill = hf
                    cell.font = hfont
                    cell.alignment = center_align
                    ws.column_dimensions[cell.column_letter].width = max(16, len(str(h)) + 2)
                sample = SAMPLE_ROWS.get(t)
                if sample:
                    ws.append(sample)
                first = False
            else:
                make_sheet(wb, t, headers, SAMPLE_ROWS.get(t))
        filename = "DB_Hardware_All_Templates.xlsx"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    r = HttpResponse(buf.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    r["Content-Disposition"] = "attachment; filename=" + filename
    return r

@login_required
def employee_list(request):
    search=request.GET.get('search','')
    status_filter=request.GET.get('status','active')
    employees=Employee.objects.annotate(hw_count=Count('assigned_hardware')).order_by('first_name')
    if status_filter: employees=employees.filter(status=status_filter)
    if search:
        employees=employees.filter(Q(emp_id__icontains=search)|Q(first_name__icontains=search)|Q(last_name__icontains=search)|Q(department__icontains=search))
    return render(request,'hardware/employee_list.html',{'employees':employees,'search':search,'status_filter':status_filter})

@login_required
def employee_detail(request, pk):
    emp=get_object_or_404(Employee, pk=pk)
    assigned_hw=get_location_filtered_hardware(request.user).filter(assigned_to=emp).prefetch_related('properties')
    return render(request,'hardware/employee_detail.html',{'emp':emp,'assigned_hw':assigned_hw})

@editor_required
def employee_add(request):
    if request.method == 'POST':
        try:
            joining=request.POST.get('joining_date') or None
            emp=Employee.objects.create(
                emp_id=request.POST.get('emp_id'),
                first_name=request.POST.get('first_name',''),
                last_name=request.POST.get('last_name',''),
                email=request.POST.get('email',''),
                phone=request.POST.get('phone',''),
                state=request.POST.get('state',''),
                center_name=request.POST.get('center_name',''),
                office_type=request.POST.get('office_type',''),
                department=request.POST.get('department',''),
                designation=request.POST.get('designation',''),
                grade=request.POST.get('grade',''),
                region=request.POST.get('region',''),
                user_type=request.POST.get('user_type',''),
                location=request.POST.get('location',''),
                joining_date=joining,
            )
            # Assign hardware if selected
            hw_id = request.POST.get('assigned_hardware')
            if hw_id:
                try:
                    hw = Hardware.objects.get(pk=hw_id)
                    hw.assigned_to = emp
                    hw.save()
                except: pass
            return JsonResponse({'success':True,'id':emp.pk})
        except Exception as e:
            return JsonResponse({'success':False,'error':str(e)})
    all_locations=Hardware.objects.values_list('location',flat=True).distinct().exclude(location='').order_by('location')
    return render(request,'hardware/employee_add.html',{'all_locations':all_locations})

@editor_required
def employee_action(request, pk):
    """Fire or reactivate employee"""
    if request.method == 'POST':
        emp=get_object_or_404(Employee, pk=pk)
        action=request.POST.get('action')
        if action == 'fire':
            emp.status='fired'
            emp.save()
            # Unassign all hardware
            Hardware.objects.filter(assigned_to=emp).update(assigned_to=None)
            return JsonResponse({'success':True,'message':f'{emp.name} has been fired. All hardware unassigned.'})
        elif action == 'reactivate':
            emp.status='active'
            emp.save()
            return JsonResponse({'success':True,'message':f'{emp.name} reactivated.'})
        elif action == 'resign':
            emp.status='resigned'
            emp.save()
            Hardware.objects.filter(assigned_to=emp).update(assigned_to=None)
            return JsonResponse({'success':True,'message':f'{emp.name} marked as resigned.'})
    return JsonResponse({'success':False})

@editor_required
@editor_required
@editor_required
@editor_required
def employee_bulk_import(request):
    """Bulk import employees from Excel"""
    if request.method != "POST":
        return JsonResponse({"success": False})
    try:
        import openpyxl
    except ImportError:
        return JsonResponse({"success": False, "error": "openpyxl not installed"})
    f = request.FILES.get("excel_file")
    if not f:
        return JsonResponse({"success": False, "error": "No file uploaded"})
    try:
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active
        headers = [str(ws.cell(1, c).value or "").strip().lower() for c in range(1, ws.max_column+1)]
        added = 0; updated = 0; id_changed = []; errors = []

        def safe(val):
            if val is None: return ""
            return str(val).strip()

        def col(row, *names):
            for name in names:
                if name in headers:
                    idx = headers.index(name)
                    if idx < len(row) and row[idx] is not None:
                        return safe(row[idx])
            return ""

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row): continue
            if all(v is None or str(v).strip() == "" for v in row): continue
            try:
                emp_id = col(row, "emp id", "employee id", "empid", "id")
                if not emp_id: continue

                name_full = col(row, "name", "full name", "employee name", "emp name")
                first_name = col(row, "first name", "firstname")
                last_name = col(row, "last name", "lastname")
                if name_full and not first_name:
                    parts = name_full.split(" ", 1)
                    first_name = parts[0]
                    last_name = parts[1] if len(parts) > 1 else ""

                email    = col(row, "gmail / email", "gmail", "email", "mail", "e-mail")
                state    = col(row, "state")
                center   = col(row, "center name", "center", "asset center", "asset center name")
                office_t = col(row, "office type")
                dept     = col(row, "department", "dept")
                desig    = col(row, "designation")
                grade    = col(row, "grade")
                region   = col(row, "region (state/corp)", "region", "region state/corp")
                utype    = col(row, "user type (user/backup/tba/stock)", "user type", "usertype", "backup/tba/stock")
                location = col(row, "location", "city", "office location")

                # Match by email first, then emp_id
                existing = None
                if email:
                    existing = Employee.objects.filter(email=email).first()

                def apply_fields(emp):
                    if first_name: emp.first_name = first_name
                    if last_name:  emp.last_name = last_name
                    if email:      emp.email = email
                    if state:      emp.state = state
                    if center:     emp.center_name = center
                    if office_t:   emp.office_type = office_t
                    if dept:       emp.department = dept
                    if desig:      emp.designation = desig
                    if grade:      emp.grade = grade
                    if region:     emp.region = region
                    if utype:      emp.user_type = utype
                    if location:   emp.location = location

                if existing:
                    old_id = existing.emp_id
                    if existing.emp_id != emp_id:
                        existing.previous_emp_id = old_id
                        existing.emp_id = emp_id
                        existing.emp_id_changed = True
                        id_changed.append({"name": existing.name, "old_id": old_id, "new_id": emp_id})
                    apply_fields(existing)
                    existing.save()
                    updated += 1
                elif Employee.objects.filter(emp_id=emp_id).exists():
                    emp = Employee.objects.get(emp_id=emp_id)
                    apply_fields(emp)
                    emp.save()
                    updated += 1
                else:
                    emp = Employee(
                        emp_id=emp_id,
                        first_name=first_name or emp_id,
                        last_name=last_name,
                        email=email, state=state,
                        center_name=center, office_type=office_t,
                        department=dept, designation=desig,
                        grade=grade, region=region,
                        user_type=utype, location=location,
                    )
                    emp.save()
                    added += 1
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")

        return JsonResponse({"success": True, "added": added, "updated": updated, "id_changed": id_changed, "errors": errors[:5]})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})

def export_employee_template(request):
    """Download Excel template for bulk employee import"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse("openpyxl not installed.", status=500)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"
    headers = [
        "Emp ID", "Name", "Gmail / Email",
        "State", "Center Name", "Office Type",
        "Department", "Designation", "Grade",
        "Region (State/Corp)", "User Type (User/Backup/TBA/Stock)",
        "Location"
    ]
    hf = PatternFill("solid", fgColor="0F1729")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hf
        cell.font = Font(bold=True, color="00F5FF", name="Consolas")
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = max(20, len(h) + 2)
    ws.append(["EMP-001", "Rahul Sharma", "rahul@gmail.com", "MP", "Bhopal Press", "Press", "IT", "Developer", "L2", "State", "User", "Bhopal"])
    ws.append(["EMP-002", "Priya Patel", "priya@gmail.com", "Delhi", "Corp Office", "Office", "Editorial", "Reporter", "M1", "Corporate", "User", "Delhi"])
    ws.append(["EMP-003", "Amit Verma", "amit@gmail.com", "Rajasthan", "Jaipur Bureau", "Bureau", "IT", "Technician", "L1", "State", "Backup", "Jaipur"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    r = HttpResponse(buf.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    r["Content-Disposition"] = "attachment; filename=Employee_Import_Template.xlsx"
    return r

def export_fired_resigned_csv(request):
    """Export CSV of fired/resigned employees"""
    if not request.user.is_authenticated: return redirect('/login/')
    employees=Employee.objects.filter(status__in=['fired','resigned']).order_by('-updated_at')
    response=HttpResponse(content_type='text/csv')
    response['Content-Disposition']=f'attachment; filename="left_employees_{date.today()}.csv"'
    writer=csv.writer(response)
    writer.writerow(['Employee ID','Previous ID','First Name','Last Name','Department','Designation','Email','Phone','Location','Joining Date','Status','Last Updated'])
    for emp in employees:
        writer.writerow([emp.emp_id,emp.previous_emp_id,emp.first_name,emp.last_name,emp.department,emp.designation,emp.email,emp.phone,emp.location,emp.joining_date,emp.status,emp.updated_at.date()])
    return response

@login_required
def employee_search_api(request):
    q=request.GET.get('q','')
    if not q: return JsonResponse({'results':[]})
    employees=Employee.objects.filter(Q(emp_id__icontains=q)|Q(first_name__icontains=q)|Q(last_name__icontains=q))[:10]
    results=[]
    for emp in employees:
        hw_qs=get_location_filtered_hardware(request.user).filter(assigned_to=emp).prefetch_related('properties')
        hw_list=[{'hw_id':hw.hw_id,'type':hw.hardware_type,'brand':hw.brand,'model':hw.model_name,'serial':hw.serial_number,'status':hw.status,'specs':hw.specifications,'location':hw.location,'properties':[{'key':p.key,'value':p.value} for p in hw.properties.all()]} for hw in hw_qs]
        results.append({'id':emp.pk,'emp_id':emp.emp_id,'name':emp.name,'department':emp.department,'email':emp.email,'phone':emp.phone,'designation':emp.designation,'hardware':hw_list})
    return JsonResponse({'results':results})

# ─── RESIGNATIONS ─────────────────────────────────────────────────────────────

@admin_required
def resignation_list(request):
    resignations=ResignationRequest.objects.order_by('-submitted_at')
    if request.user.has_location_filter:
        resignations=resignations.filter(location=request.user.location)
    return render(request,'hardware/resignation_list.html',{'resignations':resignations})

@admin_required
def resignation_review(request, pk):
    res=get_object_or_404(ResignationRequest, pk=pk)
    if request.method == 'POST':
        action=request.POST.get('action')
        note=request.POST.get('note','')
        res.reviewed_by=request.user
        res.reviewed_at=timezone.now()
        res.review_note=note
        if action == 'approve':
            res.status='approved'
            res.save()
            try:
                emp=Employee.objects.get(emp_id=res.emp_id)
                emp.status='resigned'
                emp.save()
                Hardware.objects.filter(assigned_to=emp).update(assigned_to=None)
            except Employee.DoesNotExist: pass
            # Return JSON for AJAX, redirect for normal form
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success':True,'message':'Resignation approved. Employee marked as resigned.'})
            from django.shortcuts import redirect
            return redirect('/approvals/?tab=resign')
        elif action == 'decline':
            res.status='declined'
            res.save()
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success':True,'message':'Resignation declined.'})
            from django.shortcuts import redirect
            return redirect('/approvals/?tab=resign')
    return render(request,'hardware/resignation_detail.html',{'res':res})

# ─── TRASH ────────────────────────────────────────────────────────────────────

@login_required
def trash_list(request):
    return render(request,'hardware/trash_list.html',{'trash':TrashHardware.objects.order_by('-disposed_date'),'hardware_types':get_hw_types()})

@editor_required
def trash_add(request):
    if request.method == 'POST':
        try:
            t=TrashHardware.objects.create(
                hw_id=request.POST.get('hw_id'),hardware_type=request.POST.get('hardware_type'),
                brand=request.POST.get('brand'),model_name=request.POST.get('model_name'),
                serial_number=request.POST.get('serial_number'),reason=request.POST.get('reason'),
                condition=request.POST.get('condition'),original_price=request.POST.get('original_price') or None,
                notes=request.POST.get('notes',''),disposed_by=request.user,
            )
            return JsonResponse({'success':True,'id':t.pk})
        except Exception as e:
            return JsonResponse({'success':False,'error':str(e)})
    return render(request,'hardware/trash_add.html',{'hardware_types':get_hw_types()})

# ─── USERS ────────────────────────────────────────────────────────────────────

@admin_required
def user_list(request):
    users=CustomUser.objects.all().order_by('role','username')
    all_locations=Hardware.objects.values_list('location',flat=True).distinct().exclude(location='').order_by('location')
    return render(request,'hardware/user_list.html',{'users':users,'all_locations':all_locations})

@superadmin_required
def user_create(request):
    if request.method == 'POST':
        try:
            user=CustomUser.objects.create_user(
                username=request.POST.get('username'),password=request.POST.get('password'),
                first_name=request.POST.get('first_name',''),last_name=request.POST.get('last_name',''),
                email=request.POST.get('email',''),role=request.POST.get('role','viewer'),
                phone=request.POST.get('phone',''),department=request.POST.get('department',''),
                location=request.POST.get('location',''),created_by=request.user,
            )
            return JsonResponse({'success':True,'id':user.pk})
        except Exception as e:
            return JsonResponse({'success':False,'error':str(e)})
    all_locations=Hardware.objects.values_list('location',flat=True).distinct().exclude(location='').order_by('location')
    return render(request,'hardware/user_create.html',{'all_locations':all_locations})

@superadmin_required
def user_toggle(request, pk):
    user=get_object_or_404(CustomUser,pk=pk)
    if user==request.user: return JsonResponse({'success':False,'error':'Cannot deactivate yourself'})
    user.is_active=not user.is_active; user.save()
    return JsonResponse({'success':True,'active':user.is_active})

@superadmin_required
def user_delete(request, pk):
    if request.method == 'POST':
        user = get_object_or_404(CustomUser, pk=pk)
        if user == request.user:
            return JsonResponse({'success': False, 'error': 'Cannot delete yourself'})
        username = user.username
        user.delete()
        return JsonResponse({'success': True, 'message': f'User {username} deleted.'})
    return JsonResponse({'success': False})

@superadmin_required
def user_role_change(request, pk):
    if request.method=='POST':
        user=get_object_or_404(CustomUser,pk=pk)
        new_role=request.POST.get('role')
        if new_role in ('viewer','editor','admin','superadmin'):
            user.role=new_role; user.save()
            return JsonResponse({'success':True})
    return JsonResponse({'success':False})

@superadmin_required
def user_location_change(request, pk):
    if request.method=='POST':
        user=get_object_or_404(CustomUser,pk=pk)
        user.location=request.POST.get('location','').strip(); user.save()
        return JsonResponse({'success':True,'location':user.location or 'All Locations'})
    return JsonResponse({'success':False})

# ─── COMMANDS ─────────────────────────────────────────────────────────────────

@login_required
def commands_list(request):
    cmds=CommandLog.objects.all().order_by('platform','title')
    return render(request,'hardware/commands.html',{'cmds':cmds})

@editor_required
def command_add(request):
    if request.method=='POST':
        try:
            cmd=CommandLog.objects.create(
                title=request.POST.get('title'),description=request.POST.get('description',''),
                platform=request.POST.get('platform','powershell'),command=request.POST.get('command'),
                category=request.POST.get('category',''),created_by=request.user,
            )
            return JsonResponse({'success':True,'id':cmd.pk})
        except Exception as e:
            return JsonResponse({'success':False,'error':str(e)})
    return JsonResponse({'success':False})

@login_required
def command_download(request, pk):
    cmd=get_object_or_404(CommandLog,pk=pk)
    ext={'python':'py','powershell':'ps1','cmd':'bat'}.get(cmd.platform,'txt')
    r=HttpResponse(cmd.command,content_type='text/plain')
    r['Content-Disposition']=f'attachment; filename="{cmd.title.replace(" ","_").lower()}.{ext}"'
    return r

# ─── AUTO FETCH API ───────────────────────────────────────────────────────────

@csrf_exempt
def auto_fetch_api(request):
    if request.method!='POST': return JsonResponse({'success':False,'error':'POST only'})
    try: data=json.loads(request.body)
    except: return JsonResponse({'success':False,'error':'Invalid JSON'})
    user=authenticate(request,username=data.get('username'),password=data.get('password'))
    if not user: return JsonResponse({'success':False,'error':'Invalid credentials'})
    if not user.can_edit: return JsonResponse({'success':False,'error':'Insufficient role'})
    hw_data=data.get('hardware',{})
    hw_id=hw_data.get('hw_id','').strip()
    serial=hw_data.get('serial_number','').strip()
    if not hw_id or not serial: return JsonResponse({'success':False,'error':'hw_id and serial_number required'})
    if Hardware.objects.filter(serial_number=serial).exists():
        hw=Hardware.objects.get(serial_number=serial)
        hw.properties.all().delete()
        for i,(k,v) in enumerate(hw_data.get('properties',{}).items()):
            if k and v: HardwareProperty.objects.create(hardware=hw,key=k,value=str(v),order=i)
        hw.save()
        return JsonResponse({'success':True,'action':'updated','hw_id':hw.hw_id,'message':f'Hardware {hw.hw_id} updated'})
    try:
        try: price=float(hw_data.get('price',0) or 0)
        except: price=0
        pd=hw_data.get('purchase_date','')
        try: pd=datetime.strptime(pd,'%Y-%m-%d').date() if pd else date.today()
        except: pd=date.today()
        hw=Hardware.objects.create(
            hw_id=hw_id,hardware_type=hw_data.get('hardware_type','Desktop'),
            brand=hw_data.get('brand','Unknown'),model_name=hw_data.get('model_name','Unknown'),
            serial_number=serial,purchase_date=pd,price=price,status='active',
            location=hw_data.get('location',''),specifications=hw_data.get('specifications',''),
            notes=hw_data.get('notes','Auto-added via fetch script'),created_by=user,
        )
        for i,(k,v) in enumerate(hw_data.get('properties',{}).items()):
            if k and v: HardwareProperty.objects.create(hardware=hw,key=k,value=str(v),order=i)
        return JsonResponse({'success':True,'action':'created','hw_id':hw.hw_id,'message':f'Hardware {hw.hw_id} saved!'})
    except Exception as e:
        return JsonResponse({'success':False,'error':str(e)})

# ─── STATS ────────────────────────────────────────────────────────────────────

@login_required
def stats_api(request):
    hw_qs=get_location_filtered_hardware(request.user)
    hw_types=get_hw_types()
    return JsonResponse({'hw_by_type':{t['name']:hw_qs.filter(hardware_type=t['name']).count() for t in hw_types},'total':hw_qs.count(),'active':hw_qs.filter(status='active').count(),'trash':TrashHardware.objects.count(),'employees':Employee.objects.filter(status='active').count()})

def resignation_detail(request, pk):
    res = get_object_or_404(ResignationRequest, pk=pk)
    if request.method == 'POST':
        return resignation_review(request, pk)
    return render(request, 'hardware/resignation_detail.html', {'res': res})

# ─── APPROVALS & TASKS ────────────────────────────────────────────────────────

@login_required
def approvals_dashboard(request):
    """Main approvals tab — shows resignations, transfers, hardware approvals, tasks"""
    user = request.user

    # Filter by location for admins
    def loc_filter(qs, field='location'):
        if user.has_location_filter:
            return qs.filter(**{field: user.location})
        return qs

    resignations = loc_filter(ResignationRequest.objects.filter(status='pending').order_by('-submitted_at'))
    transfers = TransferRequest.objects.filter(status='pending').order_by('-created_at')
    if user.has_location_filter:
        transfers = transfers.filter(from_location=user.location)
    hw_approvals = loc_filter(HardwareApproval.objects.filter(status='pending').order_by('-submitted_at'))
    tasks = Task.objects.filter(status__in=['open','in_progress']).order_by('-created_at')
    if user.has_location_filter:
        tasks = tasks.filter(location=user.location)

    portal_users = CustomUser.objects.filter(is_active=True).order_by('username')
    context = {
        'resignations': resignations,
        'transfers': transfers,
        'hw_approvals': hw_approvals,
        'tasks': tasks,
        'portal_users': portal_users,
        'total_pending': resignations.count() + transfers.count() + hw_approvals.count(),
    }
    return render(request, 'hardware/approvals.html', context)


# ─── TRANSFER ─────────────────────────────────────────────────────────────────

@editor_required
@editor_required
def transfer_request_create(request):
    if request.method == 'POST':
        try:
            emp = get_object_or_404(Employee, pk=request.POST.get('employee_id'))
            to_loc = request.POST.get('to_location','').strip()
            reason = request.POST.get('reason','').strip()
            is_superadmin_global = request.user.is_superadmin_role and not request.user.has_location_filter

            tr = TransferRequest.objects.create(
                employee=emp,
                from_location=emp.location,
                to_location=to_loc,
                reason=reason,
                requested_by=request.user,
                status='approved' if is_superadmin_global else 'pending',
            )
            if tr.status == 'approved':
                emp.location = to_loc
                emp.save()
                Hardware.objects.filter(assigned_to=emp).update(location=to_loc)
                return JsonResponse({'success': True, 'message': f'{emp.name} transferred to {to_loc} successfully!'})
            else:
                # Notify area superadmins about pending transfer request
                msg_title = f"Transfer Request: {emp.name} ({emp.emp_id})"
                msg_body = f"Admin {request.user.username} wants to transfer {emp.name} from {emp.location} to {to_loc}.\nReason: {reason}\n\nGo to Approvals tab to Approve or Decline."
                send_message_to_area_admins(
                    location=emp.location,
                    title=msg_title,
                    body=msg_body,
                    msg_type='transfer',
                    sender_name=request.user.username,
                    related_id=tr.pk,
                )
                return JsonResponse({'success': True, 'message': f'Transfer request submitted! Superadmin will review it.'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    employees = Employee.objects.filter(status='active').order_by('first_name')
    all_locations = Hardware.objects.values_list('location', flat=True).distinct().exclude(location='').order_by('location')
    return render(request, 'hardware/transfer.html', {'employees': employees, 'all_locations': all_locations})

@admin_required
@admin_required
def transfer_review(request, pk):
    tr = get_object_or_404(TransferRequest, pk=pk)
    if request.method == 'POST':
        action = request.POST.get('action')
        note = request.POST.get('note','')
        tr.reviewed_by = request.user
        tr.reviewed_at = timezone.now()
        tr.review_note = note
        if action == 'approve':
            tr.status = 'approved'
            tr.save()
            tr.employee.location = tr.to_location
            tr.employee.save()
            Hardware.objects.filter(assigned_to=tr.employee).update(location=tr.to_location)
            if tr.requested_by:
                send_message(tr.requested_by,
                    f"Transfer Approved: {tr.employee.name}",
                    f"Transfer of {tr.employee.name} to {tr.to_location} has been APPROVED. Note: {note}",
                    msg_type='transfer', sender_name=request.user.username, related_id=tr.pk)
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': f'{tr.employee.name} transferred to {tr.to_location}'})
            from django.shortcuts import redirect
            return redirect('/approvals/?tab=transfer')
        elif action == 'decline':
            tr.status = 'declined'
            tr.save()
            if tr.requested_by:
                send_message(tr.requested_by,
                    f"Transfer Declined: {tr.employee.name}",
                    f"Transfer of {tr.employee.name} to {tr.to_location} has been DECLINED. Reason: {note}",
                    msg_type='transfer', sender_name=request.user.username, related_id=tr.pk)
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': 'Transfer declined.', 'note': note, 'declined': True})
            from django.shortcuts import redirect
            return redirect('/approvals/?tab=transfer')
    return JsonResponse({'success': False})

@csrf_exempt
def hardware_approval_submit(request):
    """Public — employee submits hardware repair/replace/new request"""
    if request.method == 'POST':
        try:
            emp_id = request.POST.get('emp_id','').strip()
            emp_name = request.POST.get('employee_name','').strip()
            hw_id_text = request.POST.get('hw_id','').strip()
            req_type = request.POST.get('request_type','repair')
            issue = request.POST.get('issue_description','').strip()

            location = ''
            try:
                emp = Employee.objects.get(emp_id=emp_id)
                location = emp.location
            except Employee.DoesNotExist:
                pass

            hw = None
            if hw_id_text:
                try: hw = Hardware.objects.get(hw_id=hw_id_text)
                except Hardware.DoesNotExist: pass

            ha = HardwareApproval.objects.create(
                emp_id=emp_id, employee_name=emp_name,
                hardware=hw, hw_id_text=hw_id_text,
                request_type=req_type, issue_description=issue,
                location=location,
            )
            send_message_to_area_admins(
                location=location,
                title=f'HW {req_type.title()} Request: {emp_name} ({emp_id})',
                body='HW ' + req_type + ' Request from ' + emp_name + ' (' + emp_id + '). HW: ' + (hw_id_text or 'New') + '. Issue: ' + issue,
                msg_type='hw_approval',
                sender_name=emp_name,
                related_id=ha.pk,
            )
            return JsonResponse({'success': True, 'message': 'Request submitted! IT team will review it.'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return render(request, 'hardware/hardware_approval_submit.html')


@admin_required
@admin_required
@admin_required
def hardware_approval_review(request, pk):
    ha = get_object_or_404(HardwareApproval, pk=pk)
    if request.method == 'POST':
        action = request.POST.get('action')
        note = request.POST.get('note', '')
        assigned_hw_id = request.POST.get('assigned_hw_id', '')

        # ── ASSIGN TO ADMIN ──────────────────────────────────────────────────
        if action == 'assign_admin':
            admin_id = request.POST.get('admin_id')
            try:
                admin_user = CustomUser.objects.get(pk=admin_id)
                send_message(
                    admin_user,
                    f"HW Request Assigned: {ha.employee_name} ({ha.emp_id})",
                    f"You have been assigned to handle this HW request.\nEmployee: {ha.employee_name} ({ha.emp_id})\nRequest Type: {ha.get_request_type_display()}\nReported HW: {ha.hw_id_text or 'N/A'}\nIssue: {ha.issue_description}\n\nGo to Approvals → HW Requests to handle it.",
                    msg_type='hw_approval',
                    sender_name=request.user.username,
                    related_id=ha.pk,
                )
                return JsonResponse({'success': True, 'message': f'Assigned to {admin_user.username}. Message sent.'})
            except CustomUser.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Admin not found'})

        # ── DECLINE ──────────────────────────────────────────────────────────
        elif action == 'decline':
            ha.status = 'declined'
            ha.reviewed_by = request.user
            ha.reviewed_at = timezone.now()
            ha.review_note = note
            ha.save()
            return JsonResponse({'success': True, 'message': 'Request declined.', 'declined': True, 'note': note})

        # ── APPROVE ──────────────────────────────────────────────────────────
        elif action == 'approve':
            ha.reviewed_by = request.user
            ha.reviewed_at = timezone.now()
            ha.review_note = note
            ha.status = 'approved'

            msg_parts = ["Hardware request approved."]

            # Step 1: Send reported/old hardware to trash
            if ha.hw_id_text:
                try:
                    old_hw = Hardware.objects.get(hw_id=ha.hw_id_text)
                    # Move to trash
                    trash_id = 'TRASH-' + old_hw.hw_id
                    if TrashHardware.objects.filter(hw_id=trash_id).exists():
                        trash_id = trash_id + '-' + str(old_hw.pk)
                    TrashHardware.objects.create(
                        hw_id=trash_id,
                        hardware_type=old_hw.hardware_type,
                        brand=old_hw.brand,
                        model_name=old_hw.model_name,
                        serial_number=old_hw.serial_number,
                        reason=f"Reported by {ha.emp_id} ({ha.employee_name}): {ha.issue_description}",
                        condition='Damaged',
                        original_price=old_hw.price,
                        notes=f"Moved to trash on HW request approval by {request.user.username}",
                        disposed_by=request.user,
                    )
                    old_hw.delete()
                    msg_parts.append(f"{ha.hw_id_text} moved to Trash.")
                except Hardware.DoesNotExist:
                    pass  # Hardware already gone or wrong ID

            # Step 2: Assign new hardware to employee
            if assigned_hw_id:
                try:
                    new_hw = Hardware.objects.get(pk=assigned_hw_id)
                    ha.assigned_hardware = new_hw
                    try:
                        emp = Employee.objects.get(emp_id=ha.emp_id)
                        new_hw.assigned_to = emp
                        new_hw.save()
                        msg_parts.append(f"{new_hw.hw_id} ({new_hw.brand} {new_hw.model_name}) assigned to {emp.name}.")
                    except Employee.DoesNotExist:
                        new_hw.save()
                        msg_parts.append(f"{new_hw.hw_id} assigned.")
                except Hardware.DoesNotExist:
                    pass

            ha.save()
            return JsonResponse({'success': True, 'message': ' '.join(msg_parts)})

    return JsonResponse({'success': False})

@admin_required
def task_list(request):
    tasks = Task.objects.all().order_by('-created_at')
    if request.user.has_location_filter:
        tasks = tasks.filter(location=request.user.location)
    users = CustomUser.objects.filter(is_active=True)
    return render(request, 'hardware/tasks.html', {'tasks': tasks, 'users': users})


@admin_required
def task_create(request):
    if request.method == 'POST':
        try:
            task = Task.objects.create(
                title=request.POST.get('title'),
                description=request.POST.get('description',''),
                priority=request.POST.get('priority','medium'),
                due_date=request.POST.get('due_date') or None,
                location=request.POST.get('location',''),
                created_by=request.user,
            )
            assigned_id = request.POST.get('assigned_to')
            if assigned_id:
                task.assigned_to_id = assigned_id
                task.save()
            return JsonResponse({'success': True, 'id': task.pk})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False})


@admin_required
def task_update(request, pk):
    if request.method == 'POST':
        task = get_object_or_404(Task, pk=pk)
        new_status = request.POST.get('status')
        if new_status in ('open','in_progress','done','cancelled'):
            task.status = new_status
            task.save()
            return JsonResponse({'success': True})
    return JsonResponse({'success': False})


# ─── TRASH SOLD PRICE ─────────────────────────────────────────────────────────

@editor_required
def trash_mark_sold(request, pk):
    if request.method == 'POST':
        t = get_object_or_404(TrashHardware, pk=pk)
        try:
            t.sold_price = request.POST.get('sold_price') or None
            t.sold_date = request.POST.get('sold_date') or None
            t.sold_to = request.POST.get('sold_to','')
            t.sold_notes = request.POST.get('sold_notes','')
            t.save()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False})


# ─── SEND TO TRASH FROM HARDWARE EDIT ─────────────────────────────────────────

@editor_required
def hardware_send_to_trash(request, pk):
    """Send hardware directly to trash from edit/detail page"""
    if request.method == 'POST':
        hw = get_object_or_404(Hardware, pk=pk)
        if not user_can_access_hw(request.user, hw):
            return JsonResponse({'success': False, 'error': 'Location restricted'})
        try:
            trash_id = 'TRASH-' + hw.hw_id
            if TrashHardware.objects.filter(hw_id=trash_id).exists():
                trash_id = trash_id + '-' + str(hw.pk)
            TrashHardware.objects.create(
                hw_id=trash_id,
                hardware_type=hw.hardware_type,
                brand=hw.brand,
                model_name=hw.model_name,
                serial_number=hw.serial_number,
                reason=request.POST.get('reason','Sent to trash'),
                condition=request.POST.get('condition','Other'),
                original_price=hw.price,
                notes=hw.notes,
                disposed_by=request.user,
            )
            hw.delete()
            return JsonResponse({'success': True, 'message': f'{hw.hw_id} moved to trash!'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False})

# ─── MESSAGES / NOTIFICATIONS ────────────────────────────────────────────────

def send_message(recipient, title, body, msg_type='general', sender_name='System', related_id=None):
    """Helper to create in-portal message"""
    try:
        Message.objects.create(
            recipient=recipient, title=title, body=body,
            message_type=msg_type, sender_name=sender_name,
            related_id=related_id,
        )
    except Exception:
        pass


def send_message_to_area_admins(location, title, body, msg_type='general', sender_name='System', related_id=None):
    """Send message to all superadmins/admins of a specific location"""
    # Superadmins with no location = see all
    global_superadmins = CustomUser.objects.filter(role='superadmin', location='', is_active=True)
    for u in global_superadmins:
        send_message(u, title, body, msg_type, sender_name, related_id)
    # Area admins/superadmins with matching location
    if location:
        area_admins = CustomUser.objects.filter(role__in=['admin','superadmin'], location=location, is_active=True)
        for u in area_admins:
            send_message(u, title, body, msg_type, sender_name, related_id)


@login_required
def messages_list(request):
    msgs = Message.objects.filter(recipient=request.user).order_by('-created_at')[:50]
    unread = Message.objects.filter(recipient=request.user, is_read=False).count()
    # Mark all as read
    Message.objects.filter(recipient=request.user, is_read=False).update(is_read=True)
    return render(request, 'hardware/messages.html', {'msgs': msgs, 'unread': unread})


@login_required
def messages_unread_count(request):
    count = Message.objects.filter(recipient=request.user, is_read=False).count()
    return JsonResponse({'count': count})

# ─── EMPLOYEE MASTER SYNC ────────────────────────────────────────────────────

@editor_required  
def employee_master_sync(request):
    """
    Monthly employee master sync:
    - Upload new employee master sheet
    - Employees not in sheet get flagged as 'unmatched'  
    - Hardware assigned to unmatched employees gets flagged
    - Shows list of unmatched assets with option to update employee ID
    """
    if request.method == 'POST':
        action = request.POST.get('action','')
        
        # ── UPDATE HARDWARE EMPLOYEE ID ──────────────────────────────────────
        if action == 'update_hw_emp':
            hw_pk = request.POST.get('hw_pk')
            new_emp_id = request.POST.get('new_emp_id','').strip()
            try:
                hw = Hardware.objects.get(pk=hw_pk)
                try:
                    emp = Employee.objects.get(emp_id=new_emp_id)
                    hw.assigned_to = emp
                    hw.save()
                    return JsonResponse({'success': True, 'message': f'Hardware {hw.hw_id} assigned to {emp.name}'})
                except Employee.DoesNotExist:
                    return JsonResponse({'success': False, 'error': f'Employee {new_emp_id} not found'})
            except Hardware.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Hardware not found'})
        
        # ── UPLOAD MASTER SHEET ───────────────────────────────────────────────
        try:
            import openpyxl
        except ImportError:
            return JsonResponse({'success': False, 'error': 'openpyxl not installed'})
        
        f = request.FILES.get('excel_file')
        if not f:
            return JsonResponse({'success': False, 'error': 'No file uploaded'})
        
        try:
            wb = openpyxl.load_workbook(f, data_only=True)
            ws = wb.active
            headers = [str(ws.cell(1, c).value or '').strip().lower() for c in range(1, ws.max_column+1)]
            
            def safe(val):
                if val is None: return ''
                return str(val).strip()
            
            def col(row, *names):
                for name in names:
                    if name in headers:
                        idx = headers.index(name)
                        if idx < len(row) and row[idx] is not None:
                            return safe(row[idx])
                return ''
            
            # Collect all emp IDs from uploaded sheet
            sheet_emp_ids = set()
            added = 0; updated = 0; id_changed = []
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row): continue
                emp_id = col(row, 'emp id', 'employee id', 'empid')
                if not emp_id: continue
                sheet_emp_ids.add(emp_id)
                
                name = col(row, 'name', 'full name', 'employee name', 'emp name')
                email = col(row, 'gmail / email', 'gmail', 'email')
                state = col(row, 'state')
                center = col(row, 'center name', 'center')
                office_t = col(row, 'office type')
                dept = col(row, 'department')
                desig = col(row, 'designation')
                grade = col(row, 'grade')
                region = col(row, 'region (state/corp)', 'region')
                utype = col(row, 'user type (user/backup/tba/stock)', 'user type')
                location = col(row, 'location')
                
                first_name = name.split(' ',1)[0] if name else emp_id
                last_name = name.split(' ',1)[1] if name and ' ' in name else ''
                
                existing = None
                if email:
                    existing = Employee.objects.filter(email=email).first()
                
                if existing:
                    old_id = existing.emp_id
                    if existing.emp_id != emp_id:
                        existing.previous_emp_id = old_id
                        existing.emp_id = emp_id
                        existing.emp_id_changed = True
                        id_changed.append({'name': existing.name, 'old_id': old_id, 'new_id': emp_id})
                    existing.first_name = first_name or existing.first_name
                    existing.last_name = last_name or existing.last_name
                    if state: existing.state = state
                    if center: existing.center_name = center
                    if office_t: existing.office_type = office_t
                    if dept: existing.department = dept
                    if desig: existing.designation = desig
                    if grade: existing.grade = grade
                    if region: existing.region = region
                    if utype: existing.user_type = utype
                    if location: existing.location = location
                    existing.status = 'active'
                    existing.save()
                    updated += 1
                elif Employee.objects.filter(emp_id=emp_id).exists():
                    emp = Employee.objects.get(emp_id=emp_id)
                    emp.status = 'active'
                    if first_name: emp.first_name = first_name
                    if last_name: emp.last_name = last_name
                    if state: emp.state = state
                    if center: emp.center_name = center
                    if office_t: emp.office_type = office_t
                    if dept: emp.department = dept
                    if desig: emp.designation = desig
                    if grade: emp.grade = grade
                    if region: emp.region = region
                    if utype: emp.user_type = utype
                    if location: emp.location = location
                    emp.save()
                    updated += 1
                else:
                    Employee.objects.create(
                        emp_id=emp_id, first_name=first_name, last_name=last_name,
                        email=email, state=state, center_name=center,
                        office_type=office_t, department=dept, designation=desig,
                        grade=grade, region=region, user_type=utype, location=location,
                    )
                    added += 1
            
            # Mark employees NOT in sheet as unmatched (inactive)
            unmatched_count = 0
            if sheet_emp_ids:
                unmatched_qs = Employee.objects.filter(status='active').exclude(emp_id__in=sheet_emp_ids)
                unmatched_count = unmatched_qs.count()
                unmatched_qs.update(status='inactive')
            
            return JsonResponse({
                'success': True,
                'added': added, 'updated': updated,
                'unmatched': unmatched_count,
                'id_changed': id_changed,
                'message': f'Sync complete! Added: {added}, Updated: {updated}, Unmatched (not in sheet): {unmatched_count}'
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    # GET - show sync page with unmatched assets
    unmatched_employees = Employee.objects.filter(status='inactive').prefetch_related('assigned_hardware')
    unmatched_hardware = Hardware.objects.filter(
        assigned_to__status='inactive'
    ).select_related('assigned_to').order_by('hardware_type')
    
    active_employees = Employee.objects.filter(status='active').order_by('emp_id')[:500]
    return render(request, 'hardware/employee_sync.html', {
        'unmatched_employees': unmatched_employees,
        'unmatched_hardware': unmatched_hardware,
        'unmatched_hw_count': unmatched_hardware.count(),
        'unmatched_emp_count': unmatched_employees.count(),
        'active_employees': active_employees,
    })

@superadmin_required
def clear_all_data(request):
    if request.method == 'POST':
        confirm = request.POST.get('confirm','')
        if confirm == 'DELETE ALL':
            from django.db import connection
            HardwareProperty.objects.all().delete()
            Hardware.objects.all().delete()
            Employee.objects.all().delete()
            TrashHardware.objects.all().delete()
            return JsonResponse({'success': True, 'message': 'All hardware and employee data cleared!'})
        return JsonResponse({'success': False, 'error': 'Type DELETE ALL to confirm'})
    return JsonResponse({'success': False})
